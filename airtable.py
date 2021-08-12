import pandas as pd
import functools
import operator
import time
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.touch_actions import TouchActions
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from fake_useragent import UserAgent

def configure_driver():


    # Add additional Options to the webdriver
    chrome_options = Options()
    ua = UserAgent()
    userAgent = ua.random                                     #THIS IS FAKE AGENT IT WILL GIVE YOU NEW AGENT EVERYTIME
    print(userAgent)
   # add the argument and make the browser Headless.
   # chrome_options.add_argument("--headless")                    if you don't want to see the display on chrome just uncomment this


    #chrome_options.add_argument(f'user-agent={userAgent}')
    #COMMENT THE LINE OF CODE BELOW IF YOU WANT NEW RANDOM AGENT EVERYTIME INSTEAD OF SAME AGENT EVERYTIME
    #chrome_options.add_argument(
    #   '--user-agent="Mozilla/5.0 (Windows NT 6.4; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2225.0 Safari/537.36"')
    # Instantiate the Webdriver: Mention the executable path of the webdriver you have downloaded
    # For linux/Mac
    #driver = webdriver.Chrome(options = chrome_options)
    # For windows
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--disable-infobars")
    #chrome_options.add_argument("start-maximized")
    #chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument("--disable-extensions")
    #chrome_options.add_argument('--proxy-server=%s' % PROXY)
    #chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    chrome_options.add_argument("--start-maximized")
    driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
    return driver

def getCourses(driver, search_keyword):

    start_time = time.time()

    kb = openpyxl.Workbook()
    # add_sheet is used to create sheet.
    sheet1 = kb.active
    print(" WORKSHEET CREATED SUCCESSFULLY!")
    print(" ")
    print(" ")
    print(" ")
    # INITIALIZING THE COLOUMN NAMES NOW
    c1 = sheet1.cell(row=1, column=1)
    c1.value = "Title"
    c2 = sheet1.cell(row=1, column=2)
    c2.value = "Link"
    c3 = sheet1.cell(row=1, column=3)
    c3.value = "MoonShot Statement"
    c4 = sheet1.cell(row=1, column=4)
    c4.value = "Location"
    c5 = sheet1.cell(row=1, column=5)
    c5.value = "Website"
    c6 = sheet1.cell(row=1, column=6)
    c6.value = "Company Profile"
    c7 = sheet1.cell(row=1, column=7)
    c7.value = "Email"
    kb.save('airtable.xlsx')
    mi = 2
    linkno=1
    allids=[]
    mainlink='https://web.archive.org/web/20180612162227/'

    excel_data_df =pd.read_excel("erowidlink.xlsx")
    cati = excel_data_df['Links'].tolist()

    driver.get("https://airtable.com/shrJKNRdiCnuPp17E/tbl39gJXPmEeRCmK5?backgroundColor=red&viewControls=on")
    WebDriverWait(driver,80).until(expected_conditions.visibility_of_element_located((By.XPATH,'//*[@id="thirdContainer"]/div/div[1]')))
    driver.find_element_by_xpath('//*[@id="thirdContainer"]/div/div[1]').click()

    for bl in range(100):
        WebDriverWait(driver, 80).until(
            expected_conditions.visibility_of_element_located((By.XPATH, '//*[@id="galleryView"]/div/div/div/div[1]')))
        maingallery = driver.find_elements_by_xpath("//div[@class='galleryCardContainer z1']")
        print("Total: ", len(maingallery))
        for mg in maingallery:
            mainlink = mg.find_element_by_xpath(
                './/a[@class="galleryCardPrimaryCell strong truncate text-dark flex-auto line-height-4 stretched-link z1"]')
            alink = mainlink.get_attribute('href')
            title = mainlink.text
            alltitles = mg.find_elements_by_xpath('.//div[@class="px2 pt1 pb-half"]')
            moonshotstatement = ''
            Location = ''
            Website = ''
            CompanyProfile = ''
            Email = ''
            for tt in alltitles:
                heading = tt.find_element_by_xpath(
                    './/div[@class="mb-half flex items-center text-dark"]').get_attribute('title')
                if heading == 'Moonshot Statement':
                    moonshotstatement = tt.find_element_by_xpath('.//div[@class="cellContainer relative"]').text
                elif heading == 'Location':
                    Location = tt.find_element_by_xpath('.//div[@class="cellContainer relative"]').text
                elif heading == 'Website':
                    Website = tt.find_element_by_xpath('.//div[@class="cellContainer relative"]').text
                elif heading == 'Company Profile':
                    CompanyProfile = tt.find_element_by_xpath('.//div[@class="cellContainer relative"]').text
                elif heading == 'Email':
                    Email = tt.find_element_by_xpath('.//div[@class="cellContainer relative"]').text
                else:
                    pass

            print("Title: ", title)
            print("Link: ", alink)
            print("Moonshot Statement: ", moonshotstatement)
            print("Location: ", Location)
            print("Website: ", Website)
            print("Company Profile: ", CompanyProfile)
            print("Email: ", Email)
            c1 = sheet1.cell(row=mi, column=1)
            c1.value = title
            c2 = sheet1.cell(row=mi, column=2)
            c2.value = alink
            c3 = sheet1.cell(row=mi, column=3)
            c3.value = moonshotstatement
            c4 = sheet1.cell(row=mi, column=4)
            c4.value = Location
            c5 = sheet1.cell(row=mi, column=5)
            c5.value = Website
            c6 = sheet1.cell(row=mi, column=6)
            c6.value = CompanyProfile
            c7 = sheet1.cell(row=mi, column=7)
            c7.value = Email
            kb.save('airtable.xlsx')
            mi = mi+1
            print("        ")
            print("*****************************************")
            print("        ")
        print("Sleeping Do Now")
        time.sleep(7)

    print("time elapsed: {:.2f}s".format(time.time() - start_time))

# create the driver object.
search_keyword = "Web Scraping"
driver= configure_driver()
getCourses(driver, search_keyword)

# close the driver














